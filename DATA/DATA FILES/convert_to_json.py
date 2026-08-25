r"""
HHpro - Excel to JSON Converter
================================

Converts product data Excel files in this folder to JSON files that the
HHpro website uses to populate the mechanical schedules, filters, and
documentation for each product type.

Location (in the site's folder structure):
    HHpro\DATA\DATA FILES\convert_to_json.py

How it runs:
    1. Looks for every .xlsx file in the folder this script lives in.
    2. Matches each file against the PRODUCT_CONFIGS dictionary below.
    3. Reads the SCHEDULE tab and SCHEDULE NOTES tab from each matched file.
    4. Writes a .json file (same base name) into ..\JSON\ (i.e.
       HHpro\DATA\JSON\).

What this script DOES NOT do:
    - It does not check whether the filenames referenced in the Excel
      sheets actually exist on disk in ASSETS/. That check lives in the
      separate `validate_files.py` script next to this one - run it any
      time you want to know which filenames in your JSONs don't match
      real files in your ASSETS folders.

How to add a new product type:
    1. Drop the new .xlsx file into HHpro\DATA\DATA FILES\.
    2. Add one entry to PRODUCT_CONFIGS below (filename, productType,
       outputFileName, headerRows, dataStartRow, supportsMultiRow,
       assetsFolder, searchSchema).
    3. Run this script.

How to add a new DOCUMENTATION column:
    1. Add the column to row 2 of the DOCUMENTATION range in the Excel
       file(s). Use a name that starts with one of the prefixes in
       DOC_COLUMN_MAP below (e.g. "WIRING DIAGRAM" would need a new
       entry added to DOC_COLUMN_MAP first).
    2. If the column name uses a prefix not yet in DOC_COLUMN_MAP, add
       it here - key it to the ASSETS subfolder the files should live
       in and the expected file extension.
"""

import json
import os
import re
import sys
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


# -----------------------------------------------------------------------------
# PRODUCT CONFIGS
# -----------------------------------------------------------------------------
# One entry per Excel data file. The key is the input filename (exact, case
# sensitive, including '.xlsx').
#
# Fields:
#   productType        Human-readable product name for the site.
#   outputFileName     Name of the JSON file this produces (no folder path).
#   headerRows         Number of header rows below row 1 but above the data.
#                      e.g. GAS PACKS has schedule-title in row 2 and column
#                      headers in rows 3-4 -> headerRows = 3.
#   dataStartRow       1-based row where the first data row begins.
#   supportsMultiRow   True if one selection can span multiple rows (detected
#                      via merged cells). Currently only mini-splits.
#   assetsFolder       Name of this product's folder under HHpro/ASSETS/.
#                      Used by validate_files.py to check filenames against
#                      real files on disk.
#   searchSchema       Drives the Design Search page on the site. Two parts:
#                        displayName  - shown in the category picker
#                        description  - one-line blurb shown above the form
#                        targets      - list of numeric columns the engineer
#                                       can search by (target value +/- a
#                                       tolerance %). Each target = dict with
#                                       label, col (Excel column letter),
#                                       unit (display string), and
#                                       defaultTolerance (percent prefilled
#                                       in the form). Pass an empty list for
#                                       products where filter dropdowns alone
#                                       are enough (e.g. VFDs - engineers
#                                       size them to an exact HP).
# -----------------------------------------------------------------------------

PRODUCT_CONFIGS = {
    "GAS PACKS DATA.xlsx": {
        "productType": "GAS PACK RTUS",
        "outputFileName": "gas_packs.json",
        "headerRows": 3,
        "dataStartRow": 5,
        "supportsMultiRow": False,
        "assetsFolder": "GAS PACKS",
        "searchSchema": {
            "displayName": "Gas Pack RTUs",
            "description": "Packaged rooftop units. Enter design loads and the page returns models that meet the targets within your tolerance.",
            "targets": [
                {"label": "Nominal Tons",             "col": "C", "unit": "tons",  "defaultTolerance": 10},
                {"label": "Total Cooling Capacity",   "col": "G", "unit": "BTU/h", "defaultTolerance": 10},
                {"label": "Sensible Cooling Capacity","col": "H", "unit": "BTU/h", "defaultTolerance": 10},
                {"label": "Heating Output",           "col": "O", "unit": "MBH",   "defaultTolerance": 10},
                {"label": "Airflow",                  "col": "D", "unit": "CFM",   "defaultTolerance": 10},
            ],
        },
    },
    "MARVAIR DATA.xlsx": {
        "productType": "MARVAIR VERTICAL WALL MOUNT",
        "outputFileName": "marvair.json",
        "headerRows": 2,
        "dataStartRow": 4,
        "supportsMultiRow": False,
        "assetsFolder": "MARVAIR",
        "searchSchema": {
            "displayName": "Marvair Vertical Wall Mount",
            "description": "Vertical wall-mount packaged air handling units. Enter design loads and the page returns models that meet the targets within your tolerance.",
            "targets": [
                {"label": "Nominal Size",             "col": "C", "unit": "tons",  "defaultTolerance": 10},
                {"label": "Total Cooling Capacity",   "col": "E", "unit": "BTU/h", "defaultTolerance": 10},
                {"label": "Sensible Cooling Capacity","col": "F", "unit": "BTU/h", "defaultTolerance": 10},
                {"label": "Electric Heat",            "col": "I", "unit": "kW",    "defaultTolerance": 10},
                {"label": "Airflow",                  "col": "D", "unit": "CFM",   "defaultTolerance": 10},
            ],
        },
    },
    "MINI SPLIT DATA.xlsx": {
        "productType": "MINI SPLITS",
        "outputFileName": "mini_splits.json",
        "headerRows": 5,
        "dataStartRow": 7,
        "supportsMultiRow": True,
        "assetsFolder": "MINI SPLITS",
        "searchSchema": {
            "displayName": "Mini Splits",
            "description": "Ductless split systems (1:1 and multi-zone). Enter per-zone capacity targets; results include systems with at least one indoor unit matching.",
            "targets": [
                {"label": "Indoor Unit Cooling Capacity",             "col": "D", "unit": "BTU/h", "defaultTolerance": 10},
                {"label": "Indoor Unit Sensible Capacity",            "col": "E", "unit": "BTU/h", "defaultTolerance": 10},
                {"label": "Indoor Unit Heating Capacity (Heat Pump)", "col": "G", "unit": "BTU/h", "defaultTolerance": 10},
                {"label": "Indoor Unit Airflow",                      "col": "A", "unit": "CFM",   "defaultTolerance": 15},
            ],
        },
    },
    "MULTI POSITION SPLIT DATA.xlsx": {
        "productType": "MULTI POSITION SPLITS",
        "outputFileName": "multi_position_splits.json",
        "headerRows": 4,
        "dataStartRow": 6,
        "supportsMultiRow": False,
        "assetsFolder": "MULTI POSITION SPLITS",
        "searchSchema": {
            "displayName": "Multi Position Splits",
            "description": "Conventional split systems with a multi-position air handler + outdoor condensing unit. Enter design loads and the page returns models that meet the targets within your tolerance.",
            "targets": [
                {"label": "Indoor Cooling Capacity",    "col": "I", "unit": "BTU/h", "defaultTolerance": 10},
                {"label": "Indoor Sensible Capacity",   "col": "J", "unit": "BTU/h", "defaultTolerance": 10},
                {"label": "Heat Pump Heating Capacity", "col": "U", "unit": "BTU/h", "defaultTolerance": 10},
                {"label": "Aux. Electric Heat",         "col": "L", "unit": "kW",    "defaultTolerance": 10},
                {"label": "Indoor Airflow",             "col": "C", "unit": "CFM",   "defaultTolerance": 10},
            ],
        },
    },
    "GAS SPLIT DATA.xlsx": {
        "productType": "GAS SPLITS",
        "outputFileName": "gas_splits.json",
        "headerRows": 4,
        "dataStartRow": 6,
        "supportsMultiRow": False,
        "assetsFolder": "GAS SPLITS",
        "searchSchema": {
            "displayName": "Gas Splits",
            "description": "Three-component split systems with a gas furnace, indoor coil, and outdoor condensing unit. Enter design loads and the page returns models that meet the targets within your tolerance.",
            "targets": [
                {"label": "Total Cooling Capacity", "col": "H", "unit": "BTU/h",  "defaultTolerance": 10},
                {"label": "Gas Heating Output",     "col": "J", "unit": "BTU/h",  "defaultTolerance": 10},
                {"label": "Indoor Airflow",         "col": "C", "unit": "CFM",    "defaultTolerance": 10},
                {"label": "AFUE",                   "col": "L", "unit": "%",      "defaultTolerance": 5},
                {"label": "Compressor Stages",      "col": "Z", "unit": "stages", "defaultTolerance": 0},
            ],
        },
    },
    "GPS DATA.xlsx": {
        "productType": "BIPOLAR IONIZATION",
        "outputFileName": "gps.json",
        # The GPS SCHEDULE tab is unlike every other product file: it
        # stacks SEVEN independent schedules (each with its own title
        # row, header row, data rows, and NOTES block directly under
        # the data) on the one tab. multiSchedule routes the file to
        # convert_multi_schedule_file below, which emits a JSON with a
        # subSchedules list the site renders one-at-a-time (photo
        # gallery first, then the chosen product type's full schedule).
        "multiSchedule": True,
        "headerRows": 1,
        "dataStartRow": 6,
        "supportsMultiRow": False,
        "assetsFolder": "GPS",
        # Kept for shape-consistency with the other products; the site
        # excludes GPS from Design Search (mixed per-schedule columns
        # make tolerance targets meaningless for ionizers).
        "searchSchema": {
            "displayName": "Bipolar Ionization",
            "description": "GPS air ionization devices.",
            "targets": [],
        },
    },
    "VFD DATA.xlsx": {
        "productType": "VFDs",
        "outputFileName": "vfds.json",
        "headerRows": 3,
        "dataStartRow": 5,
        "supportsMultiRow": False,
        "assetsFolder": "VFDs",
        "searchSchema": {
            "displayName": "VFDs",
            "description": "Variable frequency drives. Engineers typically size VFDs to an exact motor HP and electrical service, so this category uses filters only -- no tolerance-based targets apply.",
            "targets": [],
        },
    },
    "PRICE DIFFUSER DATA.xlsx": {
        "productType": "DIFFUSERS",
        "outputFileName": "diffusers.json",
        "headerRows": 2,
        "dataStartRow": 4,
        "supportsMultiRow": False,
        "assetsFolder": "DIFFUSERS",
        # The diffuser SCHEDULE NOTES tab has TWO columns: col A is the
        # model list the note applies to ("ALL" or a comma-separated
        # list like "SPD, SCD"), col B is the note text. The site uses
        # this mapping to show only the notes that apply to the models
        # actually selected, and to auto-number each schedule row's
        # applicable notes in its Accessories column.
        "notesFormat": "modelmap",
        "searchSchema": {
            "displayName": "Diffusers",
            "description": "Price ceiling diffusers (supply + return). Enter a target airflow and/or use the filters to narrow by model, size, and application.",
            "targets": [
                {"label": "Airflow", "col": "U", "unit": "CFM", "defaultTolerance": 15},
            ],
        },
    },
    "PRICE GRILLE DATA.xlsx": {
        "productType": "GRILLES",
        "outputFileName": "grilles.json",
        "headerRows": 2,
        "dataStartRow": 4,
        "supportsMultiRow": False,
        "assetsFolder": "GRILLES",
        # Same two-column SCHEDULE NOTES mapping as the diffusers (col A
        # is "ALL" or a comma-separated MODEL list, col B is the note).
        "notesFormat": "modelmap",
        # The full grille JSON is ~27 MB and Cloudflare Pages caps files
        # at 25 MiB, so split the selections across two files
        # (grilles.json + grilles-2.json); the site re-joins them.
        "splitParts": 2,
        "searchSchema": {
            "displayName": "Grilles",
            "description": "Price supply, return, and transfer grilles. Enter a target airflow and/or use the filters to narrow by model, size, and application.",
            "targets": [
                {"label": "Airflow", "col": "I", "unit": "CFM", "defaultTolerance": 15},
            ],
        },
    },
}


# -----------------------------------------------------------------------------
# MULTI POSITION SPLIT CAPACITY TABLES
# -----------------------------------------------------------------------------
# A separate workbook (one tab per outdoor-condenser + air-handler matchup,
# tab name "<ODU> - <AHU>") that the site uses to drive the cooling +
# heat-pump capacity dropdowns on the Multi Position Split schedule. Each
# tab holds:
#   - a cooling table (cols A-F) in "long" form: every combination of
#     EAT-DB(A) / OA-cooling(B) / Airflow(C) / EAT-WB(D) has several rows
#     keyed by a VALUE type in col E, with the number in col F (RESULT).
#     We keep only MBh (total), S/T (sensible ratio) and the delta-T row
#     and ignore everything else (kW, Amps, Hi/Lo PR, etc.).
#   - an optional heat-pump table (cols H-I): outdoor ambient -> capacity
#     (MBH). Cooling-only systems omit it.
# A combination whose RESULT is "-" (or that is absent entirely) is
# invalid; we simply leave it out of the lookup, so the site only ever
# offers valid combinations in the dropdowns.
# -----------------------------------------------------------------------------

CAPACITY_FILE = "Multi Position Split Capacity Tables.xlsx"
CAPACITY_OUTPUT = "multi_position_split_capacity.json"
# Only these VALUE types (col E) feed the schedule; everything else is junk.
CAPACITY_VALUE_TYPES = ("MBh", "S/T", "∆T")  # ∆ == the delta sign


def _num_key(v):
    """Stringify a numeric cell for use as a JSON lookup key (e.g. 70.0 ->
    '70', -5 -> '-5'), so keys built here match String(value) on the site."""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v)


_HP_TEXT_NUM = re.compile(r"-?\d+(?:\.\d+)?")


def _scale_hp_text(text):
    """Scale the MBH numbers inside a TEXT heat-pump capacity cell to BTU/h.

    Some heat pumps list two capacities per ambient, e.g.
    "15.9 (standard), 22.5 (boost)". The site shows the cell text as-is,
    so keep the wording and just scale each number x1000:
    -> "15900 (standard), 22500 (boost)".

    Returns None when the text holds no numbers at all (e.g. "-"), so the
    row is omitted exactly like an invalid numeric row.
    """
    s = str(text).strip()
    if not _HP_TEXT_NUM.search(s):
        return None
    return _HP_TEXT_NUM.sub(
        lambda m: str(int(round(float(m.group()) * 1000))), s)


def convert_capacity_tables(input_path, output_path):
    """Convert the capacity-tables workbook to a JSON keyed by matchup.

    Output shape:
      {
        "matchups": {
          "<ODU> - <AHU>": {
            "axes":    {"eatDb":[...], "eatWb":[...], "oaCooling":[...], "airflow":[...]},
            "cooling": {"<eatDb>|<eatWb>|<oaCooling>|<airflow>": {"ct":N,"cs":N,"lat":N}, ...},
            "hpAxis":  [65, 60, ... -5],     # only when a heat-pump table exists
            "hp":      {"47": 17400, ...}    # outdoor ambient -> total cap (BTU/h);
                                             # boosted units hold a text value, e.g.
                                             # "15900 (standard), 22500 (boost)"
          }, ...
        }
      }
    Only VALID cooling combos are emitted (invalid/"-"/missing are omitted),
    so the site can constrain its dropdowns to valid combinations.
    """
    print(f"\nConverting: {os.path.basename(input_path)}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    matchups = {}
    skipped_tabs = []
    for sheet_name in wb.sheetnames:
        name = sheet_name.strip()
        if " - " not in name:
            skipped_tabs.append(sheet_name)
            continue
        ws = wb[sheet_name]

        combos = {}        # (db, wb, oa, cfm) -> {"MBh":.., "S/T":.., delta:..}
        axes = {"eatDb": set(), "eatWb": set(), "oaCooling": set(), "airflow": set()}
        hp = {}

        for r in range(2, ws.max_row + 1):
            # Cooling table (A-F)
            etype = ws.cell(row=r, column=5).value
            if etype is not None and str(etype).strip() in CAPACITY_VALUE_TYPES:
                db = ws.cell(row=r, column=1).value
                oa = ws.cell(row=r, column=2).value
                cfm = ws.cell(row=r, column=3).value
                wbv = ws.cell(row=r, column=4).value
                if None not in (db, oa, cfm, wbv):
                    combos.setdefault((db, oa, cfm, wbv), {})[str(etype).strip()] = \
                        ws.cell(row=r, column=6).value
                    axes["eatDb"].add(db)
                    axes["oaCooling"].add(oa)
                    axes["airflow"].add(cfm)
                    axes["eatWb"].add(wbv)
            # Heat-pump table (H-I). Most cells are a plain MBH number;
            # boosted heat pumps use a text cell listing both modes
            # ("15.9 (standard), 22.5 (boost)") which is scaled to BTU/h
            # but kept as text (see _scale_hp_text).
            amb = ws.cell(row=r, column=8).value
            cap = ws.cell(row=r, column=9).value
            if isinstance(amb, (int, float)):
                if isinstance(cap, (int, float)):
                    hp[_num_key(amb)] = int(round(cap * 1000))
                elif isinstance(cap, str) and str(cap).strip().upper() != "MBH":
                    scaled = _scale_hp_text(cap)
                    if scaled is not None:
                        hp[_num_key(amb)] = scaled

        # Keep only fully-valid cooling combos.
        cooling = {}
        for (db, oa, cfm, wbv), vals in combos.items():
            mbh = vals.get("MBh")
            st = vals.get("S/T")
            dt = vals.get("∆T")
            if not all(isinstance(x, (int, float)) for x in (mbh, st, dt)):
                continue  # invalid ("-") or incomplete -> omit
            key = f"{_num_key(db)}|{_num_key(wbv)}|{_num_key(oa)}|{_num_key(cfm)}"
            ct = int(round(mbh * 1000))
            # Sensible: round the MBH product to 3 decimals, then x1000.
            cs = int(round(round(mbh * st, 3) * 1000))
            lat_raw = db - dt
            lat = int(lat_raw) if float(lat_raw).is_integer() else round(lat_raw, 2)
            # Stored as a compact [total, sensible, LAT] triple (see _meta).
            cooling[key] = [ct, cs, lat]

        entry = {
            "axes": {k: sorted(v) for k, v in axes.items()},
            "cooling": cooling,
        }
        if hp:
            entry["hpAxis"] = sorted(hp.keys(), key=lambda x: float(x), reverse=True)
            entry["hp"] = hp
        matchups[name] = entry

    payload = {
        "matchups": matchups,
        "_meta": {
            "sourceFile":    os.path.basename(input_path),
            "generatedAt":   datetime.now().isoformat(timespec="seconds"),
            "matchupCount":  len(matchups),
            "coolingKey":    "eatDb|eatWb|oaCooling|airflow",
            "coolingValue":  "[total BTU/h, sensible BTU/h, LAT degF]",
            "hp":            "outdoor ambient (degF) -> heat-pump total BTU/h "
                             "(text when standard/boost modes are listed)",
        },
    }
    # Pure lookup data (not hand-edited) - written compact to keep the
    # file the site fetches small.
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, separators=(",", ":"), ensure_ascii=False)

    hp_count = sum(1 for m in matchups.values() if "hp" in m)
    print(f"  -> {len(matchups)} matchups written to {os.path.basename(output_path)} "
          f"({hp_count} with a heat-pump table)")
    if skipped_tabs:
        print(f"     (skipped {len(skipped_tabs)} non-matchup tab(s): "
              f"{', '.join(skipped_tabs)})")


# -----------------------------------------------------------------------------
# GAS PACK RTU CAPACITY TABLES
# -----------------------------------------------------------------------------
# "Daikin LC RTU Capacity Tables.xlsx" - four sheets covering the DSG
# (standard efficiency) and DHG (high efficiency) packaged rooftops, extracted
# from Daikin's spec sheets. It drives the condition-aware Gas Pack RTU section
# of Design Search, which is why the output is keyed by CABINET (DSG036,
# DHG090, ...) rather than by schedule model number: a cooling table is
# published per cabinet and applies to every voltage and motor built on it.
#
#   Cooling         long form, one row per EAT-DB / EAT-WB / OA-DB / CFM combo
#                   -> MBh, S/T, delta-T. Rows whose model name carries
#                   "(70% - Low Stage)" are the two-stage units' low-stage
#                   table and are stored separately under "lowStage".
#   Gas Heating     one row per cabinet + heat size (Low/Medium/High).
#   Electrical Data one row per full model number (cabinet + voltage digit +
#                   motor code), holding MCA/MOP for all four
#                   convenience-outlet / power-exhaust combinations plus the
#                   indoor motor HP.
#
# Only 208/230-3-60 (voltage digit 3) and 460-3-60 (digit 4) are carried; the
# workbook never held 575V or single phase. Combinations printed "-" in the
# spec sheet - and the three cells Daikin misprinted as 240.0, which are blank
# in the workbook - are simply omitted, so the site only ever sees rated data.
# -----------------------------------------------------------------------------

GAS_PACK_CAPACITY_FILE = "Daikin LC RTU Capacity Tables.xlsx"
GAS_PACK_CAPACITY_OUTPUT = "gas_pack_capacity.json"

# Model-number voltage digit -> the VOLT/PH spelling used by GAS PACKS DATA.
GAS_PACK_VOLTAGES = {"3": "208/3", "4": "460/3"}
# Cabinet prefix -> the Efficiency filter value on the Design Search form.
# DSG is Daikin's standard-efficiency line, DHG the high-efficiency one.
GAS_PACK_EFFICIENCY = {"DSG": "LOW", "DHG": "HIGH"}

_RANGE_RE = re.compile(r"^\s*(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)\s*$")


def _gp_first_num(v):
    """First number out of an electrical cell.

    208/230V rows print both ratings as "19.9/19.9" (208V then 230V) and some
    FLA cells carry a parenthetical variant, e.g. "2.2/1.9 (1.7/1.5)". The
    schedule shows a single 208V figure, so keep the leading number and drop
    the rest. 460V rows are already scalar and pass straight through.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).split("(")[0].split("/")[0].strip()
    try:
        return float(s)
    except ValueError:
        return None


def _gp_range(v):
    """Temperature-rise range "25-55" -> [25, 55]; anything else -> None."""
    if v is None:
        return None
    m = _RANGE_RE.match(str(v))
    if not m:
        return None
    return [float(m.group(1)), float(m.group(2))]


def _gp_trim(x):
    """Drop a pointless trailing .0 so JSON numbers stay compact."""
    return int(x) if isinstance(x, float) and x.is_integer() else x


def _gp_headers(ws):
    """Header label -> column index, so column letters can move freely."""
    out = {}
    for i, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            out[str(cell.value).strip()] = i
    return out


def _gp_cabinet(model):
    """'DSG036*D' / 'DHG036* (70% - Low Stage)' / 'DSG0363D' -> 'DSG036'."""
    return str(model).split("*")[0].strip()[:6]


def convert_gas_pack_capacity(input_path, output_path):
    """Convert the LC RTU capacity workbook to a cabinet-keyed JSON.

    Output shape:
      {
        "cabinets": {
          "DSG036": {
            "family": "DSG", "efficiency": "LOW", "tons": 3,
            "axes":    {"eatDb":[...], "eatWb":[...], "oaCooling":[...], "airflow":[...]},
            "cooling": {"<eatDb>|<eatWb>|<oaCooling>|<airflow>": [total, sensible, LAT], ...},
            "lowStage": {"axes": {...}, "cooling": {...}},   # two-stage cabinets only
            "heat": [{"size": "Low", "inputHigh": 45, "outputHigh": 36.45,
                      "inputLow": 33.75, "outputLow": 27.38,
                      "riseHigh": [15,45], "riseLow": [5,35], "thermalEff": 80}, ...],
            "electrical": {"208/3": {"D": {...}, "W": {...}}, "460/3": {...}}
          }, ...
        }
      }
    Capacities are BTU/h (the workbook is MBh) to match the schedule columns.
    LAT is EAT-DB minus the published delta-T.
    """
    print(f"\nConverting: {os.path.basename(input_path)}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    cabinets = {}

    def cab(name):
        e = cabinets.get(name)
        if e is None:
            e = cabinets[name] = {
                "family": name[:3],
                "efficiency": GAS_PACK_EFFICIENCY.get(name[:3], "LOW"),
                "tons": None,
                "axes": {"eatDb": set(), "eatWb": set(),
                         "oaCooling": set(), "airflow": set()},
                "cooling": {},
                "_lowAxes": {"eatDb": set(), "eatWb": set(),
                             "oaCooling": set(), "airflow": set()},
                "_lowCooling": {},
                "heat": [],
                "electrical": {},
            }
        return e

    # ----- Cooling -------------------------------------------------------
    ws = wb["Cooling"]
    h = _gp_headers(ws)
    need = ["Model", "Tons", "CFM", "EAT DB (°F)", "EAT WB (°F)",
            "OA DB (°F)", "MBh", "S/T", "∆T (°F)"]
    missing = [n for n in need if n not in h]
    if missing:
        raise ValueError(f"Cooling sheet is missing column(s): {missing}")

    skipped_cells = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        model = row[h["Model"] - 1]
        if not model:
            continue
        name = _gp_cabinet(model)
        low_stage = "70%" in str(model)
        entry = cab(name)
        entry["tons"] = row[h["Tons"] - 1]

        mbh = row[h["MBh"] - 1]
        st = row[h["S/T"] - 1]
        dt = row[h["∆T (°F)"] - 1]
        if not all(isinstance(x, (int, float)) for x in (mbh, st, dt)):
            skipped_cells += 1      # blank misprint cell - never invent a value
            continue

        db = row[h["EAT DB (°F)"] - 1]
        wbt = row[h["EAT WB (°F)"] - 1]
        oa = row[h["OA DB (°F)"] - 1]
        cfm = row[h["CFM"] - 1]

        axes = entry["_lowAxes"] if low_stage else entry["axes"]
        table = entry["_lowCooling"] if low_stage else entry["cooling"]
        axes["eatDb"].add(db)
        axes["eatWb"].add(wbt)
        axes["oaCooling"].add(oa)
        axes["airflow"].add(cfm)

        key = "|".join(_num_key(v) for v in (db, wbt, oa, cfm))
        lat_raw = db - dt
        table[key] = [
            int(round(mbh * 1000)),
            int(round(round(mbh * st, 3) * 1000)),
            int(lat_raw) if float(lat_raw).is_integer() else round(lat_raw, 2),
        ]

    # ----- Gas heating ---------------------------------------------------
    ws = wb["Gas Heating"]
    h = _gp_headers(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        model = row[h["Model"] - 1]
        if not model:
            continue
        entry = cab(str(model).strip())
        if entry["tons"] is None:
            entry["tons"] = row[h["Tons"] - 1]
        entry["heat"].append({
            "size": row[h["Gas Heat"] - 1],
            "inputHigh": _gp_trim(row[h["Input - High Stage (MBH)"] - 1]),
            "outputHigh": _gp_trim(row[h["Output - High Stage (MBH)"] - 1]),
            "inputLow": _gp_trim(row[h["Input - Low Stage (MBH)"] - 1]),
            "outputLow": _gp_trim(row[h["Output - Low Stage (MBH)"] - 1]),
            "riseHigh": _gp_range(row[h["Temp Rise - High Stage (°F)"] - 1]),
            "riseLow": _gp_range(row[h["Temp Rise - Low Stage (°F)"] - 1]),
            "thermalEff": _gp_trim(row[h["Thermal Efficiency (%)"] - 1]),
        })

    # ----- Electrical ----------------------------------------------------
    ws = wb["Electrical Data"]
    h = _gp_headers(ws)
    no_hp = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        model = row[h["Model"] - 1]
        if not model:
            continue
        model = str(model).strip()
        volt = GAS_PACK_VOLTAGES.get(model[6:7])
        motor = model[7:8]
        if not volt or not motor:
            continue
        entry = cab(_gp_cabinet(model))
        if entry["tons"] is None:
            entry["tons"] = row[h["Tons"] - 1]
        hp = row[h["Indoor Motor HP"] - 1] if "Indoor Motor HP" in h else None
        if hp is None:
            no_hp.append(model)
        entry["electrical"].setdefault(volt, {})[motor] = {
            "model": model,
            "mca": _gp_first_num(row[h["MCA - Base Unit"] - 1]),
            "mop": _gp_first_num(row[h["MOP - Base Unit"] - 1]),
            "hp": _gp_trim(hp) if hp is not None else None,
            "convFla": _gp_first_num(row[h["Conv. Outlet FLA"] - 1]),
            "peFla": _gp_first_num(row[h["Power Exhaust FLA"] - 1]),
            "mcaConv": _gp_first_num(row[h["MCA - w/ Conv. Outlet"] - 1]),
            "mopConv": _gp_first_num(row[h["MOP - w/ Conv. Outlet"] - 1]),
            "mcaPe": _gp_first_num(row[h["MCA - w/ Power Exhaust"] - 1]),
            "mopPe": _gp_first_num(row[h["MOP - w/ Power Exhaust"] - 1]),
            "mcaBoth": _gp_first_num(row[h["MCA - w/ Conv. Outlet + Pwr Exhaust"] - 1]),
            "mopBoth": _gp_first_num(row[h["MOP - w/ Conv. Outlet + Pwr Exhaust"] - 1]),
        }

    # ----- Finalise ------------------------------------------------------
    no_cooling = []
    for name, entry in cabinets.items():
        entry["axes"] = {k: sorted(v) for k, v in entry["axes"].items()}
        low_axes = {k: sorted(v) for k, v in entry.pop("_lowAxes").items()}
        low_cooling = entry.pop("_lowCooling")
        if low_cooling:
            entry["lowStage"] = {"axes": low_axes, "cooling": low_cooling}
        entry["motors"] = sorted({m for v in entry["electrical"].values() for m in v})
        if not entry["cooling"]:
            # DSG150 today: SS-DSG7-R32 reprints the DSG120 tables under the
            # DSG150 heading, so no real 12.5-ton standard-efficiency cooling
            # data exists. Keep the cabinet (its heating and electrical are
            # sound) and flag it so the UI can say why rather than go quiet.
            entry["coolingUnavailable"] = True
            no_cooling.append(name)

    payload = {
        "cabinets": cabinets,
        "_meta": {
            "sourceFile": os.path.basename(input_path),
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "cabinetCount": len(cabinets),
            "coolingKey": "eatDb|eatWb|oaCooling|airflow",
            "coolingValue": "[total BTU/h, sensible BTU/h, LAT degF]",
            "heat": "MBH input/output per stage; rise ranges as [min, max] degF",
            "electrical": "208V figure of each 208/230V pair; MCA/MOP per "
                          "convenience-outlet / power-exhaust combination",
        },
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, separators=(",", ":"), ensure_ascii=False)

    cool_n = sum(len(e["cooling"]) for e in cabinets.values())
    low_n = sum(len(e["lowStage"]["cooling"]) for e in cabinets.values()
                if "lowStage" in e)
    print(f"  -> {len(cabinets)} cabinets written to {os.path.basename(output_path)} "
          f"({cool_n} cooling points, {low_n} low-stage, "
          f"{sum(len(e['heat']) for e in cabinets.values())} heat sizes)")
    if skipped_cells:
        print(f"     ({skipped_cells} cooling row(s) skipped - blank/misprinted "
              f"in the source spec sheet)")
    if no_cooling:
        print(f"     (no cooling data: {', '.join(sorted(no_cooling))})")
    if no_hp:
        print(f"     (no indoor motor HP: {', '.join(sorted(no_hp))})")


# -----------------------------------------------------------------------------
# DOCUMENTATION COLUMN -> FOLDER / EXTENSION MAP
# -----------------------------------------------------------------------------
# Maps the column-name prefix (e.g. "SUBMITTAL (SYSTEM)" -> "SUBMITTAL") to
# the ASSETS subfolder name and the file extension used for that doc type.
#
# Every product's ASSETS folder has this same flat structure of 8
# sub-folders, so the same map applies to every product:
#
#   HHpro\ASSETS\<product>\SUBMITTALS\           (.pdf)
#   HHpro\ASSETS\<product>\ENGINEERING MANUALS\  (.pdf)
#   HHpro\ASSETS\<product>\CAPACITY TABLES\      (.pdf)
#   HHpro\ASSETS\<product>\INSTALLATION MANUALS\ (.pdf)
#   HHpro\ASSETS\<product>\REVIT\                (.zip)
#   HHpro\ASSETS\<product>\CAD\                  (.zip)
#   HHpro\ASSETS\<product>\CONTROLS\             (.pdf)
#   HHpro\ASSETS\<product>\OPERATION MANUAL\     (.pdf)
#
# Order matters: longer / more specific prefixes must come BEFORE shorter
# ones so that e.g. "OPERATION MANUAL" is matched before any hypothetical
# shorter prefix.
# -----------------------------------------------------------------------------

DOC_COLUMN_MAP = [
    ("ENGINEERING MANUAL",  {"folder": "ENGINEERING MANUALS",  "fileExtension": "pdf"}),
    ("INSTALLATION MANUAL", {"folder": "INSTALLATION MANUALS", "fileExtension": "pdf"}),
    ("OPERATION MANUAL",    {"folder": "OPERATION MANUAL",     "fileExtension": "pdf"}),
    ("CAPACITY TABLE",      {"folder": "CAPACITY TABLES",      "fileExtension": "pdf"}),
    ("CONTROLS OPTIONS",    {"folder": "CONTROLS",             "fileExtension": "pdf"}),
    ("SUBMITTAL",           {"folder": "SUBMITTALS",           "fileExtension": "pdf"}),
    ("REVIT",               {"folder": "REVIT",                "fileExtension": "zip"}),
    ("CAD",                 {"folder": "CAD",                  "fileExtension": "zip"}),
]


def resolve_doc_folder(column_name):
    """Map a documentation column name (e.g. "SUBMITTAL (SYSTEM)") to its
    target folder and file extension via DOC_COLUMN_MAP.

    Returns a dict {folder, fileExtension}. If nothing matches, returns a
    dict with a generic guess and prints a warning (so a new-but-unmapped
    column gets noticed rather than silently dropped).
    """
    upper = (column_name or "").strip().upper()
    for prefix, info in DOC_COLUMN_MAP:
        if upper.startswith(prefix):
            return {"folder": info["folder"], "fileExtension": info["fileExtension"]}

    # Unknown column - synthesize a folder name from the column itself so
    # the JSON still produces a reachable URL, and flag it to the user.
    fallback = re.sub(r"\s*\([^)]*\)\s*", "", column_name or "").strip().upper()
    print(f"  WARNING: Documentation column '{column_name}' doesn't match "
          f"any prefix in DOC_COLUMN_MAP. Using '{fallback}' as the folder; "
          f"add an entry to DOC_COLUMN_MAP if this is a new doc type.")
    return {"folder": fallback or "UNKNOWN", "fileExtension": "pdf"}


# -----------------------------------------------------------------------------
# HELPERS
# -----------------------------------------------------------------------------

def clean_value(val):
    """Normalize a cell value for JSON output."""
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip()
        if v == "":
            return None
        return v
    if isinstance(val, float):
        # Round to drop float noise like 15.965000000000001 while keeping
        # precision when it's meaningful.
        if val.is_integer():
            return int(val)
        return round(val, 6)
    if isinstance(val, datetime):
        return val.isoformat()
    return val


def find_row1_sections(ws):
    """
    Find the row-1 merged sections and return their column ranges as
    1-based (start, end) tuples.

    Required sections (every product file must have these):
      SCHEDULE RANGE, FILTERS, DOCUMENTATION

    Optional sections (only on products that need them):
      REFRIGERANT CALCULATIONS  - mini-splits / multi-position splits
                                  for the line-set + charge calculator
                                  on the project view.
    """
    REQUIRED = ("SCHEDULE RANGE", "FILTERS", "DOCUMENTATION")
    OPTIONAL = ("REFRIGERANT CALCULATIONS",)
    KNOWN = REQUIRED + OPTIONAL

    sections = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row == 1:
            label = clean_value(ws.cell(row=1, column=mr.min_col).value)
            if label in KNOWN:
                sections[label] = (mr.min_col, mr.max_col)

    missing = [k for k in REQUIRED if k not in sections]
    if missing:
        raise ValueError(
            "Row 1 is missing required merged section(s): " + ", ".join(missing) +
            ". Each of SCHEDULE RANGE, FILTERS, DOCUMENTATION must be a "
            "single merged cell in row 1 spanning its columns."
        )
    return sections


def merge_lookup(ws):
    """Build a (row, col) -> MergedCellRange map for the whole sheet plus
    a set of anchor (row, col) pairs (top-left of each merge).

    Lets us ask in O(1):
      - "is this cell covered by a merge?"
      - "is this cell the anchor of a merge?"
    """
    cell_to_merge = {}
    anchors = set()
    for mr in ws.merged_cells.ranges:
        anchors.add((mr.min_row, mr.min_col))
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                cell_to_merge[(r, c)] = mr
    return cell_to_merge, anchors


def extract_schedule_header(ws, schedule_cols, header_rows, cell_to_merge, anchors):
    """Read the schedule header rows (rows 2 .. 2+header_rows-1) inside the
    schedule range, preserving any horizontal and vertical merges.

    Returns a list (one element per header row) of lists of
    {col, value, colspan, rowspan} dicts. Only the anchor cell of a merge
    is emitted; cells covered by a merge are skipped.
    """
    min_col, max_col = schedule_cols
    rows_out = []
    for r in range(2, 2 + header_rows):
        row_cells = []
        for c in range(min_col, max_col + 1):
            mr = cell_to_merge.get((r, c))
            if mr is not None and (r, c) != (mr.min_row, mr.min_col):
                continue  # covered by a merge, not the anchor

            val = clean_value(ws.cell(row=r, column=c).value)
            colspan = 1
            rowspan = 1
            if mr is not None:
                colspan = mr.max_col - mr.min_col + 1
                rowspan = mr.max_row - mr.min_row + 1

            row_cells.append({
                "col":     get_column_letter(c),
                "value":   val,
                "colspan": colspan,
                "rowspan": rowspan,
            })
        rows_out.append(row_cells)
    return rows_out


def extract_filter_columns(ws, filter_cols):
    """Read the FILTER columns' names from row 2."""
    min_col, max_col = filter_cols
    out = []
    for c in range(min_col, max_col + 1):
        name = clean_value(ws.cell(row=2, column=c).value)
        if name:
            out.append({"name": name, "col": get_column_letter(c)})
    return out


def extract_doc_columns(ws, doc_cols):
    """Read the DOCUMENTATION columns' names from row 2 and map each one
    to its ASSETS sub-folder + file extension via DOC_COLUMN_MAP.
    """
    min_col, max_col = doc_cols
    docs = []
    for c in range(min_col, max_col + 1):
        name = clean_value(ws.cell(row=2, column=c).value)
        if not name:
            continue
        mapping = resolve_doc_folder(name)
        docs.append({
            "name":          name,
            "col":           get_column_letter(c),
            "folder":        mapping["folder"],
            "fileExtension": mapping["fileExtension"],
        })
    return docs


def extract_refrigerant_columns(ws, refrigerant_cols):
    """Read the REFRIGERANT CALCULATIONS columns' header names from row 2.

    Headers in this section follow the same row-2 convention used by
    FILTERS and DOCUMENTATION: each column has a single cell on row 2
    containing the full label (newlines inside the label are flattened
    so the JSON name is a single line).
    """
    min_col, max_col = refrigerant_cols
    out = []
    for c in range(min_col, max_col + 1):
        name = clean_value(ws.cell(row=2, column=c).value)
        if not name:
            continue
        # Excel cells can hold a soft-wrapped label like "MAX VERTICAL\nSEPARATION\n(ODU TO IDU)\n(FT)";
        # squash any embedded newlines + collapse whitespace so the JSON name
        # reads cleanly in the site UI.
        flat = re.sub(r"\s+", " ", str(name)).strip()
        out.append({"name": flat, "col": get_column_letter(c)})
    return out


def find_last_data_row(ws, data_start_row, schedule_cols):
    """
    Scan down from data_start_row and return the last row with any value
    in the schedule range. Trims trailing empty rows (some files pad to
    row 1,048,570).
    """
    min_col, max_col = schedule_cols
    last = data_start_row - 1
    empty_streak = 0
    row = data_start_row
    while row <= ws.max_row:
        has_value = False
        for c in range(min_col, max_col + 1):
            if ws.cell(row=row, column=c).value not in (None, ""):
                has_value = True
                break
        if has_value:
            last = row
            empty_streak = 0
        else:
            empty_streak += 1
            # Safety: stop after 20 consecutive empty rows
            if empty_streak >= 20:
                break
        row += 1
    return last


def build_groups(ws, data_start_row, data_end_row, schedule_cols,
                 supports_multi_row, cell_to_merge):
    """Group consecutive data rows into selections.

    For products without multi-row selections, each data row is its own
    group. For multi-row products, rows linked by a multi-row merge in
    the schedule range are joined into one group.
    """
    if not supports_multi_row:
        return [(r, r) for r in range(data_start_row, data_end_row + 1)]

    min_col, max_col = schedule_cols
    row_groups = {r: [r, r] for r in range(data_start_row, data_end_row + 1)}

    for mr in ws.merged_cells.ranges:
        if mr.min_col < min_col or mr.max_col > max_col:
            continue
        if mr.min_row < data_start_row or mr.max_row > data_end_row:
            continue
        if mr.max_row == mr.min_row:
            continue
        for r in range(mr.min_row, mr.max_row + 1):
            bounds = row_groups[r]
            bounds[0] = min(bounds[0], mr.min_row)
            bounds[1] = max(bounds[1], mr.max_row)

    groups = []
    r = data_start_row
    while r <= data_end_row:
        lo, hi = row_groups[r]
        changed = True
        while changed:
            changed = False
            for rr in range(lo, hi + 1):
                b = row_groups[rr]
                if b[0] < lo:
                    lo = b[0]; changed = True
                if b[1] > hi:
                    hi = b[1]; changed = True
        groups.append((lo, hi))
        r = hi + 1
    return groups


def row_data_for_range(ws, row, col_start, col_end):
    """Read one row of data between col_start and col_end. Keys are
    column letters; only cells with a value are included."""
    data = {}
    for c in range(col_start, col_end + 1):
        val = clean_value(ws.cell(row=row, column=c).value)
        if val is None:
            continue
        data[get_column_letter(c)] = val
    return data


def row_data_for_range_by_name(ws, row, columns_meta):
    """Same as row_data_for_range but keyed by the column's display name
    instead of its letter. Used for filterData and documentationData."""
    out = {}
    for meta in columns_meta:
        c = column_index_from_string(meta["col"])
        val = clean_value(ws.cell(row=row, column=c).value)
        if val is None:
            continue
        out[meta["name"]] = val
    return out


def row_horizontal_spans(cell_to_merge, row, col_start, col_end):
    """Find horizontal (colspan > 1) merges on a single data row within
    the schedule range. Returns {anchor_col_letter: colspan}.

    Example: on a Mini Splits row where J:L is merged (Voltage/MCA/MOP)
    with "Indoor Powered From Outdoor Unit", returns {"J": 3}.

    Vertical-only merges (rowspan > 1, colspan == 1) are ignored - those
    are already handled by the row-group logic that builds multi-row
    selections.
    """
    spans = {}
    seen_merges = set()
    for c in range(col_start, col_end + 1):
        mr = cell_to_merge.get((row, c))
        if mr is None:
            continue
        if id(mr) in seen_merges:
            continue
        seen_merges.add(id(mr))
        colspan = mr.max_col - mr.min_col + 1
        if colspan <= 1:
            continue
        if mr.min_col < col_start or mr.max_col > col_end:
            continue
        if mr.min_row != row:
            continue
        spans[get_column_letter(mr.min_col)] = colspan
    return spans


def extract_selections(ws, groups, schedule_cols, filter_cols, doc_cols,
                       filter_columns_meta, doc_columns_meta, cell_to_merge,
                       refrigerant_columns_meta=None):
    """Build the selections list from the grouped row ranges.

    `refrigerant_columns_meta` is optional (only present on Mini Splits
    and Multi Position Splits today). When provided, each row gets a
    `refrigerantData` field keyed by the column's display name -- same
    shape as filterData / documentationData.
    """
    selections = []
    for i, (lo, hi) in enumerate(groups, start=1):
        sel_id = f"sel_{i:04d}"
        rows_out = []
        for r in range(lo, hi + 1):
            row_entry = {
                "scheduleData":      row_data_for_range(ws, r, schedule_cols[0], schedule_cols[1]),
                "filterData":        row_data_for_range_by_name(ws, r, filter_columns_meta),
                "documentationData": row_data_for_range_by_name(ws, r, doc_columns_meta),
            }
            if refrigerant_columns_meta:
                row_entry["refrigerantData"] = row_data_for_range_by_name(
                    ws, r, refrigerant_columns_meta
                )
            spans = row_horizontal_spans(cell_to_merge, r, schedule_cols[0], schedule_cols[1])
            if spans:
                row_entry["scheduleCellSpans"] = spans
            rows_out.append(row_entry)
        selections.append({"id": sel_id, "rows": rows_out})
    return selections


# -----------------------------------------------------------------------------
# SCHEDULE NOTES
# -----------------------------------------------------------------------------

def extract_schedule_notes(wb, notes_format=None):
    """Read the SCHEDULE NOTES tab and return a structured notes object.

    `notes_format` comes from the product config:
      "modelmap"  - two-column layout (col A = model list or "ALL",
                    col B = note text). Used by DIFFUSERS.
      None        - auto-detect: the Marvair three-section layout is
                    recognized by cell A1 starting with 'STANDARD
                    OPTIONS'; otherwise a flat one-note-per-row list.

    Output shape:
      {"format": "list",     "notes": [...]}
      {"format": "modelmap", "notes": [{"models": [...], "text": ...}, ...]}
      {"format": "marvair",  "standard": [...], "configuration": [...],
                             "optional":  [{"text": ..., "sub": [...]}, ...]}
    """
    if "SCHEDULE NOTES" not in wb.sheetnames:
        return {"format": "list", "notes": []}
    ws = wb["SCHEDULE NOTES"]

    if notes_format == "modelmap":
        return _parse_modelmap_notes(ws)

    a1 = ws.cell(row=1, column=1).value
    a1_str = str(a1).strip().upper() if a1 is not None else ""
    if a1_str.startswith("STANDARD OPTIONS"):
        return _parse_marvair_notes(ws)
    return _parse_simple_notes(ws)


def _parse_modelmap_notes(ws):
    """Parse the two-column model-mapped SCHEDULE NOTES layout.

    Col A: "ALL" or a comma-separated model list ("SPD, SCD, SCDA").
    Col B: the note text. Rows missing either cell are skipped.
    Note order in the sheet is preserved - it drives the numbering
    on the site.
    """
    notes = []
    for r in range(1, ws.max_row + 1):
        models_raw = clean_value(ws.cell(row=r, column=1).value)
        text = clean_value(ws.cell(row=r, column=2).value)
        if not models_raw or not text:
            continue
        models = [m.strip() for m in str(models_raw).split(",") if m.strip()]
        notes.append({"models": models, "text": str(text)})
    return {"format": "modelmap", "notes": notes}


def _parse_simple_notes(ws):
    """Read notes from column A of the SCHEDULE NOTES tab, one per row.
    Blank rows are skipped."""
    notes = []
    for r in range(1, ws.max_row + 1):
        val = clean_value(ws.cell(row=r, column=1).value)
        if val:
            notes.append(str(val))
    return {"format": "list", "notes": notes}


def _parse_marvair_notes(ws):
    """Parse the three-section Marvair SCHEDULE NOTES layout.

    Expected structure:
      Row 1:      A = 'STANDARD OPTIONS/ACCESSORIES:'
      Rows 2..N:  Left notes   - A = '1-' / B = text
                  Right notes  - on the same rows, some further column
                                  has 'N-' followed by text
      Row M:      A = 'OPTIONAL ACCESSORIES:'
      Some row:   some column has 'CONFIGURATION:'
      Below OPTIONAL:
        - Main notes:  A = 'N-' / B = text
        - Sub-notes:   B = '-' / C = text (tied to the main note above)
      Below CONFIGURATION: numbered notes in config_col / config_col+1.
    """
    standard = []
    configuration = []
    optional = []

    # Locate the three section markers
    standard_row = None
    optional_row = None
    config_row = None
    config_col = None

    for r in range(1, ws.max_row + 1):
        a_raw = ws.cell(row=r, column=1).value
        if a_raw is not None:
            a_upper = str(a_raw).strip().upper()
            if standard_row is None and a_upper.startswith("STANDARD OPTIONS"):
                standard_row = r
            elif optional_row is None and a_upper.startswith("OPTIONAL ACCESSORIES"):
                optional_row = r
        if config_row is None:
            max_scan = min(ws.max_column + 1, 30)
            for c in range(1, max_scan):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                v = str(val).strip().upper()
                if v == "CONFIGURATION:" or v.startswith("CONFIGURATION:"):
                    config_row = r
                    config_col = c
                    break

    # STANDARD: left column first, then right column, so output order is
    # 1..14 rather than interleaved.
    std_start = (standard_row + 1) if standard_row else 1
    std_end = (optional_row - 1) if optional_row else ws.max_row

    for r in range(std_start, std_end + 1):
        b = ws.cell(row=r, column=2).value
        if b is not None and str(b).strip():
            standard.append(str(b).strip())

    for r in range(std_start, std_end + 1):
        for c in range(3, min(ws.max_column + 1, 20)):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            if re.match(r"^\d+-$", str(val).strip()):
                text_cell = ws.cell(row=r, column=c + 1).value
                if text_cell is not None and str(text_cell).strip():
                    standard.append(str(text_cell).strip())
                break

    # CONFIGURATION
    if config_row is not None and config_col is not None:
        for r in range(config_row + 1, ws.max_row + 1):
            num = ws.cell(row=r, column=config_col).value
            if num is None:
                continue
            if not re.match(r"^\d+-$", str(num).strip()):
                break
            text = ws.cell(row=r, column=config_col + 1).value
            if text is not None and str(text).strip():
                configuration.append(str(text).strip())

    # OPTIONAL
    if optional_row is not None:
        current_parent = None
        for r in range(optional_row + 1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value

            if a is not None and re.match(r"^\d+-$", str(a).strip()):
                text = str(b).strip() if (b is not None and str(b).strip()) else ""
                current_parent = {"text": text, "sub": []}
                optional.append(current_parent)
            elif (b is not None and str(b).strip() == "-"
                    and c is not None and str(c).strip()):
                if current_parent is not None:
                    current_parent["sub"].append(str(c).strip())

    return {
        "format":        "marvair",
        "standard":      standard,
        "configuration": configuration,
        "optional":      optional,
    }


# -----------------------------------------------------------------------------
# MULTI-SCHEDULE CONVERSION (GPS)
# -----------------------------------------------------------------------------
# The GPS DATA file stacks several independent schedules on the one
# SCHEDULE tab. Each block looks like:
#
#   <title row>     merged across the schedule width (A:L or A:K)
#   <header row>    col A = "ZONE TAG"; cells merged vertically over
#                   1-2 spacer rows below
#   <data rows>     one selection per row
#   NOTES:          marker row
#   <note lines>    one per row in col A, pre-numbered in the text
#                   ("1. Install...", "-- optional specifications --", ...)
#
# Row 1 still carries the SCHEDULE RANGE / FILTERS / DOCUMENTATION
# sections, but FILTERS is a single (unmerged) cell, so this parser
# resolves the sections itself instead of using find_row1_sections.
#
# Output: the standard payload fields (filterColumns, documentation-
# Columns, selections with globally-unique ids) PLUS a subSchedules
# list: [{index, title, filterValue, photoKey, scheduleHeader,
# selectionIds, notes}]. Note text keeps its own numbering verbatim -
# the data rows' NOTES column cites those numbers, so the site and the
# exports must never renumber them.
# -----------------------------------------------------------------------------

def _ms_clean_text(val):
    """Note/title text: normalize NBSP + collapse runs of whitespace."""
    if val is None:
        return None
    s = str(val).replace("\u00a0", " ")
    s = re.sub(r"[ \t]+", " ", s).strip()
    return s or None


def _ms_row1_sections(ws):
    """Resolve the row-1 section column ranges, allowing single-cell
    (unmerged) sections. Returns {label: (min_col, max_col)}."""
    KNOWN = ("SCHEDULE RANGE", "FILTERS", "DOCUMENTATION")
    sections = {}
    merged = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row == 1:
            merged[mr.min_col] = (mr.min_col, mr.max_col)
    for c in range(1, ws.max_column + 1):
        label = clean_value(ws.cell(row=1, column=c).value)
        if label in KNOWN and label not in sections:
            sections[label] = merged.get(c, (c, c))
    missing = [k for k in KNOWN if k not in sections]
    if missing:
        raise ValueError("Row 1 is missing section label(s): " + ", ".join(missing))
    return sections


def convert_multi_schedule_file(input_path, config, output_path):
    """Convert a stacked-multi-schedule Excel file (GPS) to JSON."""
    print(f"\nConverting: {os.path.basename(input_path)}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    if "SCHEDULE" not in wb.sheetnames:
        raise ValueError(f"'{os.path.basename(input_path)}' has no SCHEDULE tab.")
    ws = wb["SCHEDULE"]

    sections = _ms_row1_sections(ws)
    schedule_cols = sections["SCHEDULE RANGE"]
    filter_cols = sections["FILTERS"]
    doc_cols = sections["DOCUMENTATION"]

    filter_columns = extract_filter_columns(ws, filter_cols)
    doc_columns = extract_doc_columns(ws, doc_cols)

    cell_to_merge, _anchors = merge_lookup(ws)

    def cellv(r, c):
        return clean_value(ws.cell(row=r, column=c).value)

    # --- Locate every schedule block: a non-empty row whose NEXT row
    # --- starts with "ZONE TAG" in column A is a schedule title row.
    title_rows = []
    for r in range(2, ws.max_row + 1):
        if cellv(r, 1) is None:
            continue
        below = cellv(r + 1, 1)
        if below is not None and str(below).strip().upper() == "ZONE TAG":
            title_rows.append(r)
    if not title_rows:
        raise ValueError("No schedule blocks found (title + ZONE TAG header rows).")

    sub_schedules = []
    selections = []
    sel_num = 0

    for title_row in title_rows:
        title = _ms_clean_text(ws.cell(row=title_row, column=1).value)

        # Schedule width: the title row's merged span (falls back to the
        # row-1 SCHEDULE RANGE width for an unmerged title).
        mr = cell_to_merge.get((title_row, 1))
        width = (mr.max_col if mr is not None else schedule_cols[1])

        # Header row + how far its vertical merges extend (spacer rows).
        header_row = title_row + 1
        header_end = header_row
        for c in range(1, width + 1):
            hm = cell_to_merge.get((header_row, c))
            if hm is not None and hm.max_row > header_end:
                header_end = hm.max_row

        column_letters = [get_column_letter(c) for c in range(1, width + 1)]
        header_cells = []
        for c in range(1, width + 1):
            header_cells.append({
                "col":     get_column_letter(c),
                "value":   cellv(header_row, c),
                "colspan": 1,
                "rowspan": 1,
            })

        # Data rows: everything between the header block and the NOTES:
        # marker (or the first fully-blank row, as a safety stop).
        sel_ids = []
        row = header_end + 1
        notes_marker_row = None
        while row <= ws.max_row:
            a_val = cellv(row, 1)
            if a_val is not None and str(a_val).strip().upper().startswith("NOTES"):
                notes_marker_row = row
                break
            has_value = any(cellv(row, c) is not None for c in range(1, width + 1))
            if not has_value:
                break
            sel_num += 1
            sel_id = f"sel_{sel_num:04d}"
            row_entry = {
                "scheduleData":      row_data_for_range(ws, row, 1, width),
                "filterData":        row_data_for_range_by_name(ws, row, filter_columns),
                "documentationData": row_data_for_range_by_name(ws, row, doc_columns),
            }
            selections.append({"id": sel_id, "rows": [row_entry]})
            sel_ids.append(sel_id)
            row += 1

        # Notes: one line per row in column A until the first blank row.
        notes = []
        if notes_marker_row is not None:
            r = notes_marker_row + 1
            while r <= ws.max_row:
                text = _ms_clean_text(ws.cell(row=r, column=1).value)
                if text is None:
                    break
                notes.append(text)
                r += 1

        # First row's filter value labels the gallery card; the first
        # SUBMITTAL doc value keys the card photo (DATA/PICTURES/GPS/
        # "GPS - <photoKey>.webp").
        filter_value = None
        photo_key = None
        if sel_ids:
            first = selections[len(selections) - len(sel_ids)]["rows"][0]
            for fc in filter_columns:
                v = first["filterData"].get(fc["name"])
                if v is not None:
                    filter_value = v
                    break
            doc_data = first["documentationData"]
            for dc in doc_columns:
                if dc["name"].upper().startswith("SUBMITTAL") and doc_data.get(dc["name"]):
                    photo_key = str(doc_data[dc["name"]])
                    break
            if photo_key is None:
                for dc in doc_columns:
                    if doc_data.get(dc["name"]):
                        photo_key = str(doc_data[dc["name"]])
                        break

        sub_schedules.append({
            "index":       len(sub_schedules),
            "title":       title,
            "filterValue": filter_value,
            "photoKey":    photo_key,
            "scheduleHeader": {
                "columnLetters": column_letters,
                "rows":          [header_cells],
            },
            "selectionIds": sel_ids,
            # Verbatim, pre-numbered lines (data rows cite these
            # numbers) - rendered without renumbering everywhere.
            "notes":        notes,
        })

    payload = {
        "productType":                config["productType"],
        "assetsFolder":               config.get("assetsFolder"),
        "scheduleTitle":              "AIR IONIZATION DEVICE SCHEDULES",
        "supportsMultiRowSelections": False,
        "multiSchedule":              True,
        "searchSchema":               config["searchSchema"],
        "filterColumns":              filter_columns,
        "documentationColumns":       doc_columns,
        "refrigerantColumns":         [],
        "subSchedules":               sub_schedules,
        "selections":                 selections,
        # Notes live per sub-schedule (see subSchedules[].notes).
        "scheduleNotes":              {"format": "list", "notes": []},
        "_meta": {
            "sourceFile":       os.path.basename(input_path),
            "generatedAt":      datetime.now().isoformat(timespec="seconds"),
            "subScheduleCount": len(sub_schedules),
            "selectionCount":   len(selections),
        },
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print(f"  -> {len(sub_schedules)} sub-schedules / {len(selections)} selections "
          f"written to {os.path.basename(output_path)}")
    for sub in sub_schedules:
        print(f"     [{sub['index']}] {sub['title']}")
        print(f"         filter={sub['filterValue']!r} photo={sub['photoKey']!r} "
              f"rows={len(sub['selectionIds'])} notes={len(sub['notes'])}")


# -----------------------------------------------------------------------------
# MAIN CONVERSION
# -----------------------------------------------------------------------------

def convert_file(input_path, config, output_path):
    """Convert a single Excel file to JSON using the given config."""
    print(f"\nConverting: {os.path.basename(input_path)}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    if "SCHEDULE" not in wb.sheetnames:
        raise ValueError(f"'{os.path.basename(input_path)}' has no SCHEDULE tab.")
    ws = wb["SCHEDULE"]

    sections = find_row1_sections(ws)
    schedule_cols = sections["SCHEDULE RANGE"]
    filter_cols = sections["FILTERS"]
    doc_cols = sections["DOCUMENTATION"]
    refrigerant_cols = sections.get("REFRIGERANT CALCULATIONS")  # optional

    cell_to_merge, anchors = merge_lookup(ws)

    schedule_title = clean_value(ws.cell(row=2, column=schedule_cols[0]).value)

    schedule_header_rows = extract_schedule_header(
        ws, schedule_cols, config["headerRows"], cell_to_merge, anchors
    )
    filter_columns = extract_filter_columns(ws, filter_cols)
    doc_columns = extract_doc_columns(ws, doc_cols)
    refrigerant_columns = (
        extract_refrigerant_columns(ws, refrigerant_cols)
        if refrigerant_cols else []
    )

    data_end_row = find_last_data_row(ws, config["dataStartRow"], schedule_cols)
    groups = build_groups(
        ws, config["dataStartRow"], data_end_row, schedule_cols,
        config["supportsMultiRow"], cell_to_merge,
    )
    selections = extract_selections(
        ws, groups, schedule_cols, filter_cols, doc_cols,
        filter_columns, doc_columns, cell_to_merge,
        refrigerant_columns_meta=refrigerant_columns or None,
    )

    schedule_notes = extract_schedule_notes(wb, config.get("notesFormat"))

    schedule_col_letters = [
        get_column_letter(c) for c in range(schedule_cols[0], schedule_cols[1] + 1)
    ]

    payload = {
        "productType":               config["productType"],
        "assetsFolder":              config.get("assetsFolder"),
        "scheduleTitle":             schedule_title,
        "supportsMultiRowSelections": config["supportsMultiRow"],
        # searchSchema feeds the Design Search page on the site
        # (numeric "design target" inputs + filter dropdowns derived
        # from filterColumns). Required by every product config; pass
        # an empty targets list for products that need filters only.
        "searchSchema":              config["searchSchema"],
        "scheduleHeader": {
            "columnLetters": schedule_col_letters,
            "rows":          schedule_header_rows,
        },
        "filterColumns":        filter_columns,
        "documentationColumns": doc_columns,
        # refrigerantColumns is the schema (display name + Excel column
        # letter) for every column in the optional REFRIGERANT
        # CALCULATIONS section. The site reads it on the project view's
        # "Refrigerant" tab to drive the line-set + charge calculator.
        # Empty list when the product file doesn't have that section.
        "refrigerantColumns":   refrigerant_columns,
        "selections":           selections,
        "scheduleNotes":        schedule_notes,
        "_meta": {
            "sourceFile":    os.path.basename(input_path),
            "generatedAt":   datetime.now().isoformat(timespec="seconds"),
            "dataRowCount":  data_end_row - config["dataStartRow"] + 1,
            "selectionCount": len(selections),
        },
    }

    # Cloudflare Pages rejects any single file over 25 MiB. Products
    # whose JSON would exceed that (grilles: 32k+ selections) set
    # "splitParts" in PRODUCT_CONFIGS: the selections array is split
    # into that many contiguous chunks. The main file keeps everything
    # else (header, filters, docs, notes, searchSchema) plus chunk 1
    # and lists the continuation files in "continuationFiles"; each
    # continuation file holds only its chunk of selections. The site
    # (JS/data.js loadProduct) fetches the continuations and re-joins
    # the selections transparently.
    parts = int(config.get("splitParts") or 1)
    if parts > 1:
        stem, ext = os.path.splitext(output_path)
        chunk_size = (len(selections) + parts - 1) // parts
        chunks = [selections[i:i + chunk_size]
                  for i in range(0, len(selections), chunk_size)]
        cont_names = [f"{os.path.basename(stem)}-{i + 2}{ext}"
                      for i in range(len(chunks) - 1)]
        payload["selections"] = chunks[0]
        payload["continuationFiles"] = cont_names
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        for name, chunk in zip(cont_names, chunks[1:]):
            part_path = os.path.join(os.path.dirname(output_path), name)
            with open(part_path, "w", encoding="utf-8") as f:
                json.dump({"selections": chunk}, f, indent=2, ensure_ascii=False)
        sizes = ", ".join(
            f"{os.path.basename(p)} ({os.path.getsize(p) / 1048576:.1f} MiB)"
            for p in [output_path] + [
                os.path.join(os.path.dirname(output_path), n) for n in cont_names]
        )
        print(f"  -> {len(selections)} selections written across "
              f"{len(chunks)} files: {sizes}")
        return

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print(f"  -> {len(selections)} selections written to "
          f"{os.path.basename(output_path)}")
    if config["supportsMultiRow"]:
        multi = sum(1 for s in selections if len(s["rows"]) > 1)
        print(f"     ({multi} multi-row selections, "
              f"{len(selections) - multi} single-row selections)")


def choose_files(candidates):
    """Ask which of the convertible files to run this time.

    Converting everything takes a long while (the grille/diffuser files
    alone are ~30 minutes), so the script lists what it found and lets
    you convert just the file(s) you actually changed.

    Press Enter (or type "all") to convert everything; otherwise type the
    numbers of the files you want, separated by commas/spaces (e.g. "2"
    or "2, 5"). If no console input is available (e.g. run headless),
    every file is converted, matching the old behavior.
    """
    print("\nFiles available to convert:")
    for i, fname in enumerate(candidates, start=1):
        print(f"  {i}. {fname}")
    while True:
        try:
            raw = input("\nWhich file(s)? Numbers separated by commas, "
                        "or Enter for ALL: ").strip()
        except EOFError:
            print("  (no console input available - converting ALL files)")
            return candidates
        if raw == "" or raw.lower() == "all":
            return candidates
        chosen, bad = [], []
        for p in re.split(r"[,\s]+", raw):
            if not p:
                continue
            if p.isdigit() and 1 <= int(p) <= len(candidates):
                fname = candidates[int(p) - 1]
                if fname not in chosen:
                    chosen.append(fname)
            else:
                bad.append(p)
        if chosen and not bad:
            return chosen
        print(f"  Didn't understand {', '.join(bad) if bad else repr(raw)}. "
              f"Use numbers 1-{len(candidates)}, e.g. '2' or '1, 3'.")


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.abspath(os.path.join(script_dir, "..", "JSON"))
    os.makedirs(output_dir, exist_ok=True)

    print(f"Input folder : {script_dir}")
    print(f"Output folder: {output_dir}")

    all_xlsx = [f for f in sorted(os.listdir(script_dir))
                if f.lower().endswith(".xlsx") and not f.startswith("~$")]
    candidates = [f for f in all_xlsx
                  if f in (CAPACITY_FILE, GAS_PACK_CAPACITY_FILE)
                  or f in PRODUCT_CONFIGS]
    skipped = [f for f in all_xlsx if f not in candidates]

    if not candidates:
        print("\nNo convertible .xlsx files found in this folder.")
    to_convert = choose_files(candidates) if candidates else []

    converted = 0
    for fname in to_convert:
        input_path = os.path.join(script_dir, fname)
        if fname == CAPACITY_FILE:
            try:
                convert_capacity_tables(
                    input_path,
                    os.path.join(output_dir, CAPACITY_OUTPUT),
                )
                converted += 1
            except Exception as e:
                print(f"  ERROR processing {fname}: {e}")
            continue
        if fname == GAS_PACK_CAPACITY_FILE:
            try:
                convert_gas_pack_capacity(
                    input_path,
                    os.path.join(output_dir, GAS_PACK_CAPACITY_OUTPUT),
                )
                converted += 1
            except Exception as e:
                print(f"  ERROR processing {fname}: {e}")
            continue
        config = PRODUCT_CONFIGS[fname]
        output_path = os.path.join(output_dir, config["outputFileName"])
        try:
            if config.get("multiSchedule"):
                convert_multi_schedule_file(input_path, config, output_path)
            else:
                convert_file(input_path, config, output_path)
            converted += 1
        except Exception as e:
            print(f"  ERROR processing {fname}: {e}")

    print(f"\nDone. Converted {converted} of {len(to_convert)} selected file(s).")
    if skipped:
        print("Skipped (no config entry in PRODUCT_CONFIGS):")
        for f in skipped:
            print(f"  - {f}")
        print("Add an entry to PRODUCT_CONFIGS to include these files.")

    print("\nTip: run validate_files.py in this folder to check whether "
          "the filenames referenced in the generated JSONs actually exist "
          "in your ASSETS folders.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        print("\nScript stopped due to an unexpected error:")
        traceback.print_exc()
    finally:
        try:
            input("\nPress Enter to close this window...")
        except EOFError:
            pass
