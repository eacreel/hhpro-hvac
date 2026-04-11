"""
convert_mps_data.py — Convert MULTI_POSITION_SPLIT_DATA.xlsx to multi-position-splits.json

Usage:
    python convert_mps_data.py

INPUT_FILE  = "MULTI POSITION SPLIT DATA.xlsx"
OUTPUT_FILE = "JSON/multi-position-splits.json"

Schedule data:      Columns A-AE (rows 1-4 headers, row 5+ data)
Filter columns:     AG-AM (row 4 has filter names)
Document columns:   AO-BA (row 4 has document descriptions)

Filter column mapping:
    Col AG (33): SIZE
    Col AH (34): COOLING ONLY OR HEAT PUMP
    Col AI (35): NOMINAL EFFICIENCY
    Col AJ (36): ELECTRICAL
    Col AK (37): ELECTRIC HEAT KW
    Col AL (38): COMPRESSOR STAGES
    Col AM (39): FAN TYPE

Document folder mapping:
    Col AO (41): SUBMITTAL (SYSTEM)              -> SUBMITTALS/           .pdf
    Col AP (42): SUBMITTAL (OUTDOOR UNIT)        -> SUBMITTALS/           .pdf
    Col AQ (43): SUBMITTAL (INDOOR UNIT)         -> SUBMITTALS/           .pdf
    Col AR (44): ENGINEERING MANUAL (SYSTEM)     -> ENGINEERING MANUAL/   .pdf
    Col AS (45): ENGINEERING MANUAL (OUTDOOR)    -> ENGINEERING MANUAL/   .pdf
    Col AT (46): ENGINEERING MANUAL (INDOOR)     -> ENGINEERING MANUAL/   .pdf
    Col AU (47): CAPACITY TABLE                  -> CAPACITY TABLES/      .pdf
    Col AV (48): INSTALLATION MANUAL (OUTDOOR)   -> INSTALLATION MANUALS/ .pdf
    Col AW (49): INSTALLATION MANUAL (INDOOR)    -> INSTALLATION MANUALS/ .pdf
    Col AX (50): REVIT (OUTDOOR)                 -> REVIT/                .zip
    Col AY (51): REVIT (INDOOR)                  -> REVIT/                .zip
    Col AZ (52): CAD (OUTDOOR)                   -> CAD/                  .zip
    Col BA (53): CAD (INDOOR)                    -> CAD/                  .zip
"""

import openpyxl
import json
import os

# -----------------------------------------------------------------------
# Configuration
# -----------------------------------------------------------------------
INPUT_FILE  = "MULTI POSITION SPLIT DATA.xlsx"
OUTPUT_FILE = "JSON/multi-position-splits.json"
BASE_PATH   = "ASSETS/MULTI POSITION SPLITS/"

# Document column -> (subfolder, file extension)
DOC_FOLDERS = {
    41: ("SUBMITTALS/",           ".pdf"),
    42: ("SUBMITTALS/",           ".pdf"),
    43: ("SUBMITTALS/",           ".pdf"),
    44: ("ENGINEERING MANUAL/",   ".pdf"),
    45: ("ENGINEERING MANUAL/",   ".pdf"),
    46: ("ENGINEERING MANUAL/",   ".pdf"),
    47: ("CAPACITY TABLES/",      ".pdf"),
    48: ("INSTALLATION MANUALS/", ".pdf"),
    49: ("INSTALLATION MANUALS/", ".pdf"),
    50: ("REVIT/",                ".zip"),
    51: ("REVIT/",                ".zip"),
    52: ("CAD/",                  ".zip"),
    53: ("CAD/",                  ".zip"),
}

# Document column -> JSON key name
DOC_KEYS = {
    41: "submittalSystem",
    42: "submittalOutdoor",
    43: "submittalIndoor",
    44: "engineeringManualSystem",
    45: "engineeringManualOutdoor",
    46: "engineeringManualIndoor",
    47: "capacityTable",
    48: "installManualOutdoor",
    49: "installManualIndoor",
    50: "revitOutdoor",
    51: "revitIndoor",
    52: "cadOutdoor",
    53: "cadIndoor",
}

# Data starts at row 5 (rows 1-4 are headers)
DATA_START_ROW = 5


# -----------------------------------------------------------------------
# Helper Functions
# -----------------------------------------------------------------------
def clean(value):
    """Return None for empty/dash values, stripped string otherwise."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s == "-":
        return None
    return s


def to_num(value):
    """Convert to number if possible, round floats to 2 decimals."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if value == int(value):
            return int(value)
        return round(value, 2)
    s = str(value).strip()
    if s == "" or s == "-":
        return None
    try:
        f = float(s)
        if f == int(f):
            return int(f)
        return round(f, 2)
    except (ValueError, TypeError):
        return s


def to_str(value):
    """Convert to string, return None for empty/dash."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s == "-":
        return None
    return s


def to_str_rounded(value):
    """Convert to string, rounding floats to 2 decimal places."""
    if value is None:
        return None
    if isinstance(value, float):
        rounded = round(value, 2)
        if rounded == int(rounded):
            return str(int(rounded))
        return str(rounded)
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if s == "" or s == "-":
        return None
    try:
        f = float(s)
        rounded = round(f, 2)
        if rounded == int(rounded):
            return str(int(rounded))
        return str(rounded)
    except (ValueError, TypeError):
        return s


def doc_path(ws, col, row):
    """Build the full relative document path from a cell value."""
    val = clean(ws.cell(row=row, column=col).value)
    if val is None:
        return None
    folder, ext = DOC_FOLDERS[col]
    return BASE_PATH + folder + val + ext


# -----------------------------------------------------------------------
# Read Schedule Notes from "Schedule Notes" tab
# -----------------------------------------------------------------------
def read_schedule_notes(wb):
    """Read pre-made schedule notes from the 'Schedule Notes' tab."""
    notes = []
    if "Schedule Notes" in wb.sheetnames:
        ws = wb["Schedule Notes"]
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is None:
                break
            s = str(val).strip()
            if s:
                notes.append(s)
    return notes


# -----------------------------------------------------------------------
# Main Conversion
# -----------------------------------------------------------------------
def convert():
    print(f"Reading: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    # Read schedule notes from the Schedule Notes tab
    schedule_notes = read_schedule_notes(wb)
    print(f"Schedule Notes: {len(schedule_notes)} note(s)")
    for i, note in enumerate(schedule_notes):
        print(f"  {i+1}: {note}")

    # Track unique filter values for filterOptions
    size_set = set()
    sys_type_set = set()
    nom_eff_set = set()
    elec_set = set()
    heat_kw_set = set()
    comp_set = set()
    fan_set = set()

    systems = []
    sys_id = 0

    for row in range(DATA_START_ROW, ws.max_row + 1):
        # Skip empty rows (check column A)
        if ws.cell(row=row, column=1).value is None:
            continue

        sys_id += 1
        sid = "mps_{:04d}".format(sys_id)

        # ----- Indoor Air Handling Unit (Columns A-Q) -----
        indoor = {
            "symbol":               clean(ws.cell(row=row, column=1).value) or "AHU-",
            "model":                clean(ws.cell(row=row, column=2).value) or "",
            "airflow":              to_num(ws.cell(row=row, column=3).value),
            "motorHp":              to_num(ws.cell(row=row, column=4).value),
            "motorType":            to_str(ws.cell(row=row, column=5).value),
            "coolingEatDb":         to_num(ws.cell(row=row, column=6).value),
            "coolingEatWb":         to_num(ws.cell(row=row, column=7).value),
            "coolingLatDb":         to_num(ws.cell(row=row, column=8).value),
            "coolingTotal":         to_num(ws.cell(row=row, column=9).value),
            "coolingSensible":      to_num(ws.cell(row=row, column=10).value),
            "heatPumpTotalCapacity": to_num(ws.cell(row=row, column=11).value),
            "auxHeatKw":            to_str(ws.cell(row=row, column=12).value),
            "auxHeatTempRise":      to_str_rounded(ws.cell(row=row, column=13).value),
            "voltage":              to_str(ws.cell(row=row, column=14).value),
            "mca":                  to_num(ws.cell(row=row, column=15).value),
            "mop":                  to_num(ws.cell(row=row, column=16).value),
            "weight":               to_num(ws.cell(row=row, column=17).value),
        }

        # ----- Outdoor Condensing Unit (Columns R-AD) -----
        outdoor = {
            "symbol":               clean(ws.cell(row=row, column=18).value) or "CU-",
            "model":                clean(ws.cell(row=row, column=19).value) or "",
            "heatingAmbient":       to_num(ws.cell(row=row, column=20).value),
            "heatingTotal":         to_num(ws.cell(row=row, column=21).value),
            "heatingEfficiency":    to_str(ws.cell(row=row, column=22).value),
            "voltage":              to_str(ws.cell(row=row, column=23).value),
            "mca":                  to_num(ws.cell(row=row, column=24).value),
            "mop":                  to_num(ws.cell(row=row, column=25).value),
            "coolingAmbient":       to_num(ws.cell(row=row, column=26).value),
            "refrigerant":          to_str(ws.cell(row=row, column=27).value),
            "efficiency":           to_str(ws.cell(row=row, column=28).value),
            "weight":               to_num(ws.cell(row=row, column=29).value),
            "compressorStages":     to_str(ws.cell(row=row, column=30).value),
        }

        # ----- Notes (Column AE = 31) -----
        notes = clean(ws.cell(row=row, column=31).value) or ""

        # ----- Filters (Columns AG-AM = 33-39) -----
        f_size    = to_num(ws.cell(row=row, column=33).value)
        f_type    = to_str(ws.cell(row=row, column=34).value)
        f_nom_eff = to_str(ws.cell(row=row, column=35).value)
        f_elec    = to_str(ws.cell(row=row, column=36).value)
        f_heat    = to_str(ws.cell(row=row, column=37).value)
        f_comp    = to_str(ws.cell(row=row, column=38).value)
        f_fan     = to_str(ws.cell(row=row, column=39).value)

        # Collect unique filter values
        if f_size is not None:  size_set.add(f_size)
        if f_type:              sys_type_set.add(f_type)
        if f_nom_eff:           nom_eff_set.add(f_nom_eff)
        if f_elec:              elec_set.add(f_elec)
        if f_heat:              heat_kw_set.add(f_heat)
        if f_comp:              comp_set.add(f_comp)
        if f_fan:               fan_set.add(f_fan)

        filters = {
            "size":              f_size,
            "systemType":        f_type,
            "nominalEfficiency": f_nom_eff,
            "electrical":        f_elec,
            "electricHeatKw":    f_heat,
            "compressorStages":  f_comp,
            "fanType":           f_fan,
        }

        # ----- Documents (Columns AO-BA = 41-53) -----
        docs = {}
        for col_num, key in DOC_KEYS.items():
            docs[key] = doc_path(ws, col_num, row)

        # ----- Build system object -----
        system = {
            "id":           sid,
            "indoorUnit":   indoor,
            "outdoorUnit":  outdoor,
            "notes":        notes,
            "filters":      filters,
            "docs":         docs,
        }

        systems.append(system)

    # -------------------------------------------------------------------
    # Build sorted filter options
    # -------------------------------------------------------------------

    # Sizes: numeric sort
    sizes_sorted = sorted([float(x) for x in size_set])

    # System types: alphabetical
    sys_types_sorted = sorted(list(sys_type_set))

    # Nominal efficiency: alphabetical
    nom_eff_sorted = sorted(list(nom_eff_set))

    # Electrical types: alphabetical
    elec_sorted = sorted(list(elec_set))

    # Electric Heat kW: numeric sort (exclude "-" / None)
    heat_kw_clean = set()
    for v in heat_kw_set:
        if v and v != "-":
            heat_kw_clean.add(v)
    heat_kw_nums = []
    for v in heat_kw_clean:
        try:
            heat_kw_nums.append(float(v))
        except (ValueError, TypeError):
            pass
    heat_kw_sorted = [
        str(int(n)) if n == int(n) else str(n)
        for n in sorted(heat_kw_nums)
    ]

    # Compressor stages: numeric first, then text
    comp_sorted = sorted(
        list(comp_set),
        key=lambda x: (0 if x.isdigit() else 1, x)
    )

    # Fan types: alphabetical
    fan_sorted = sorted(list(fan_set))

    # -------------------------------------------------------------------
    # Assemble final JSON
    # -------------------------------------------------------------------
    data = {
        "productType":   "Multi Position Splits",
        "manufacturer":  "DAIKIN",
        "totalSystems":  len(systems),
        "filterOptions": {
            "sizes":               sizes_sorted,
            "systemTypes":         sys_types_sorted,
            "nominalEfficiencies": nom_eff_sorted,
            "electricalTypes":     elec_sorted,
            "electricHeatKw":      heat_kw_sorted,
            "compressorStages":    comp_sorted,
            "fanTypes":            fan_sorted,
        },
        "scheduleNotes": schedule_notes,
        "systems": systems,
    }

    # -------------------------------------------------------------------
    # Write output
    # -------------------------------------------------------------------
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

    with open(OUTPUT_FILE, "w") as f:
        json.dump(data, f, indent=2)

    print(f"Generated {len(systems)} systems -> {OUTPUT_FILE}")
    print(f"Filter options:")
    for key, vals in data["filterOptions"].items():
        print(f"  {key}: {vals}")


if __name__ == "__main__":
    convert()
