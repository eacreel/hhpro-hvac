"""
convert_gas_packs.py — Convert GAS PACKS DATA.xlsx to gas-packs.json

Place this script in the DATA/ folder alongside the Excel file.
Run:  python convert_gas_packs.py

Output: DATA/JSON/gas-packs.json
"""

import os
import json
import openpyxl

# ---------------------------------------------------------------------------
# Paths (resolve relative to this script's location)
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "GAS PACKS DATA.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "JSON")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "gas-packs.json")

ASSET_BASE = "ASSETS/GAS PACKS/"

# ---------------------------------------------------------------------------
# Column mappings (1-indexed)
# ---------------------------------------------------------------------------
# Schedule columns (A=1 through Y=25)
#   A(1)=TAG (skipped), B(2)=MAKE, C(3)=MODEL, ... Y(25)=NOTES (skipped)
SCHEDULE_COLS = {
    2:  "manufacturer",
    3:  "model",
    4:  "nomTons",
    5:  "cfm",
    6:  "esp",
    7:  "tesp",
    8:  "coolingTotalCapacity",
    9:  "coolingSensibleCapacity",
    10: "efficiency",
    11: "edb",
    12: "ewb",
    13: "ldb",
    14: "lwb",
    15: "heatingInput",
    16: "heatingOutput",
    17: "heatingEat",
    18: "heatingLat",
    19: "hgrh",
    20: "coolingStages",
    21: "voltage",
    22: "motorHp",
    23: "mca",
    24: "mocp",
}

# Filter columns (AA=27 through AF=32)
FILTER_COLS = {
    27: "size",
    28: "electrical",
    29: "efficiency",
    30: "coolingStages",
    31: "gasHeat",
    32: "hgrh",
}

# Document columns (AH=34 through AL=38)
DOC_COLS = {
    34: "submittal",
    35: "engineeringManual",
    36: "installationManual",
    37: "revit",
    38: "cad",
}

# Document folder names matching the ASSETS/GAS PACKS/ subfolders
DOC_FOLDERS = {
    "submittal":          "SUBMITTALS",
    "engineeringManual":  "ENGINEERING MANUALS",
    "installationManual": "INSTALLATION MANUALS",
    "revit":              "REVIT",
    "cad":                "CAD",
}

# Document file extensions
DOC_EXTENSIONS = {
    "submittal":          ".pdf",
    "engineeringManual":  ".pdf",
    "installationManual": ".pdf",
    "revit":              ".zip",
    "cad":                ".zip",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def cell_val(ws, row, col):
    """Return cell value or None."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def to_str(val):
    """Convert a value to string, keeping None as None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def smart_number(val):
    """Convert to number if possible, otherwise return string. Never parse
    slash-delimited values like '208/3' as numbers."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # Don't convert values containing slashes (e.g. "208/3")
    if "/" in s:
        return s
    try:
        f = float(s)
        # Return int if it's a whole number
        if f == int(f):
            return int(f)
        return f
    except (ValueError, TypeError):
        return s


def build_doc_path(folder_name, filename, extension):
    """Build asset path: ASSETS/GAS PACKS/<folder>/<filename><ext>"""
    if not filename:
        return None
    return ASSET_BASE + folder_name + "/" + str(filename) + extension


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------
def convert():
    if not os.path.isfile(EXCEL_PATH):
        print(f"ERROR: Cannot find '{EXCEL_PATH}'")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    systems = []
    filter_option_sets = {key: set() for key in FILTER_COLS.values()}

    # Determine data range (rows 4 to max_row)
    row = 4
    while row <= ws.max_row:
        # Stop if model column is empty (end of data)
        model = cell_val(ws, row, 3)
        if not model:
            break

        # --- Build schedule object (flat — single unit, no indoor/outdoor) ---
        schedule = {}
        for col, key in SCHEDULE_COLS.items():
            schedule[key] = smart_number(cell_val(ws, row, col))

        # --- Build filters object ---
        # All filter values stored as strings to avoid corrupting
        # slash-delimited electrical values like "208/3"
        filters = {}
        for col, key in FILTER_COLS.items():
            val = to_str(cell_val(ws, row, col))
            filters[key] = val

            # Collect unique values for filterOptions
            if val is not None:
                filter_option_sets[key].add(val)

        # --- Build docs object ---
        docs = {}
        for col, key in DOC_COLS.items():
            filename = to_str(cell_val(ws, row, col))
            folder = DOC_FOLDERS[key]
            ext = DOC_EXTENSIONS[key]
            docs[key] = build_doc_path(folder, filename, ext)

        # --- Build system entry ---
        # Append "-hgrh" suffix when HGRH is YES to avoid duplicate IDs
        # (same model number appears once with HGRH and once without)
        hgrh_suffix = "-hgrh" if filters.get("hgrh") == "YES" else ""
        system_id = "gp-" + str(model).lower() + hgrh_suffix
        system = {
            "id": system_id,
            "productKey": "gas-packs",
            "schedule": schedule,
            "filters": filters,
            "docs": docs,
        }

        systems.append(system)
        row += 1

    # --- Build filterOptions (sorted) ---
    filter_options = {}
    for key, vals in filter_option_sets.items():
        sorted_vals = sorted(vals, key=lambda v: (float(v) if v.replace(".", "", 1).isdigit() else float("inf"), v))
        filter_options[key] = sorted_vals

    # --- Assemble output ---
    output = {
        "productType": "Light Commercial RTUs - Gas Heat",
        "manufacturer": "Daikin",
        "systems": systems,
        "filterOptions": filter_options,
    }

    # --- Write JSON ---
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"Converted {len(systems)} systems → {OUTPUT_PATH}")
    print(f"Filter options: { {k: len(v) for k, v in filter_options.items()} }")


if __name__ == "__main__":
    convert()
