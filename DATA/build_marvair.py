"""
Convert MARVAIR_DATA.xlsx into marvair-vertical.json for HHpro.

USAGE (from the DATA/ folder):
    python build_marvair.py

INPUTS  (expected relative paths from this script):
    MARVAIR DATA.xlsx           — source data

OUTPUTS:
    JSON/marvair-vertical.json  — consumed by DataLoader in the web app

SOURCE SHEET LAYOUT (MARVAIR_DATA.xlsx — Sheet1):
    Rows  4–43  : 40 system data rows
    Rows 44–75  : schedule notes (section headers + numbered items)
    Cols  B–R   : schedule data (headers on rows 2–3)
    Cols  T–V   : filter values (TONS, ELECTRICAL, ELECTRIC HEAT)
    Cols  X–Y   : document filenames (DATA SHEET, IOM)

KEY DECISIONS (must stay consistent with the web app's JS):
    * productKey is "marvair-vertical"
    * Structure is flat (sys.schedule) — no indoor/outdoor split
    * Filter values stay as strings (protects "208/60/1" from parseFloat)
    * preserveNoteNumbering=True           — notes already contain "1-", "2-"
      so the web app renders them verbatim and does NOT auto-number
    * scheduleNotesDefaultChecked=True     — all notes pre-checked on first
      open of the schedule preview
    * Nominal tons kept as string ("7.5", "10", "12.5", "15")
    * System ID format: marvair_{model}_{electricHeat}kw

ASSET PATHS:
    Document paths resolve to ASSETS/MARVAIR/SUBMITTALS/... and
    ASSETS/MARVAIR/INSTALLATION MANUALS/... where the PDF basenames come
    from columns X and Y of the source data.
"""

import json
import os
import sys
import openpyxl


# ---------------------------------------------------------------------------
# Paths (relative to the script location)
# ---------------------------------------------------------------------------
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
INPUT_XLSX  = os.path.join(SCRIPT_DIR, "MARVAIR DATA.xlsx")
OUTPUT_JSON = os.path.join(SCRIPT_DIR, "JSON", "marvair-vertical.json")

ASSET_BASE     = "ASSETS/MARVAIR/"
SUBMITTAL_DIR  = "SUBMITTALS/"
IOM_DIR        = "INSTALLATION MANUALS/"


# ---------------------------------------------------------------------------
# Helpers — normalize raw Excel cell values
# ---------------------------------------------------------------------------
def to_num(val):
    """Return int if whole number, float otherwise, None if empty/non-numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return int(val)
        return val
    try:
        s = str(val).strip()
        if not s:
            return None
        n = float(s)
        if n.is_integer():
            return int(n)
        return n
    except (ValueError, TypeError):
        return val


def to_str(val):
    """Return stripped string, or empty string if None."""
    if val is None:
        return ""
    return str(val).strip()


def format_tons(val):
    """Keep nominal tons as string, preserving fractional values like 7.5."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def format_heat_int(val):
    """Electric heat comes in as a number; coerce to int for ID/filter use."""
    if isinstance(val, float) and val.is_integer():
        return int(val)
    if isinstance(val, int):
        return val
    return 0


def numkey(s):
    """Sort key — parse numeric strings, non-numeric sort last."""
    try:
        return float(s)
    except (ValueError, TypeError):
        return 9999


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------
def main():
    if not os.path.isfile(INPUT_XLSX):
        print(f"[ERROR] Input file not found: {INPUT_XLSX}")
        sys.exit(1)

    print(f"Loading {INPUT_XLSX} …")
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    ws = wb["Sheet1"]

    # -----------------------------------------------------------------------
    # Part 1 — Parse system data rows (4-43)
    # -----------------------------------------------------------------------
    systems = []
    size_set        = set()
    electrical_set  = set()
    electric_heat_set = set()

    for row_idx in range(4, 44):
        model = to_str(ws.cell(row=row_idx, column=3).value)  # Col C
        if not model:
            continue   # skip empty rows

        # Schedule data (cols B-R on sheet, but we skip B which is TAG)
        nom_tons_raw   = ws.cell(row=row_idx, column=4).value   # D — NOMINAL SIZE
        cfm            = to_num(ws.cell(row=row_idx, column=5).value)   # E
        total_cap      = to_num(ws.cell(row=row_idx, column=6).value)   # F
        sens_cap       = to_num(ws.cell(row=row_idx, column=7).value)   # G
        outside_air    = to_num(ws.cell(row=row_idx, column=8).value)   # H
        entering_air   = to_str(ws.cell(row=row_idx, column=9).value)   # I
        electric_heat  = ws.cell(row=row_idx, column=10).value          # J
        voltage        = to_str(ws.cell(row=row_idx, column=11).value)  # K
        mca            = to_num(ws.cell(row=row_idx, column=12).value)  # L
        mocp           = to_num(ws.cell(row=row_idx, column=13).value)  # M
        refrigerant    = to_str(ws.cell(row=row_idx, column=14).value)  # N
        eer            = to_num(ws.cell(row=row_idx, column=15).value)  # O
        ieer           = to_num(ws.cell(row=row_idx, column=16).value)  # P
        configuration  = to_str(ws.cell(row=row_idx, column=17).value)  # Q
        ventilation    = to_str(ws.cell(row=row_idx, column=18).value)  # R

        # Filter values — kept as strings (protects "208/60/1")
        filter_tons           = to_str(ws.cell(row=row_idx, column=20).value)  # T
        filter_electrical     = to_str(ws.cell(row=row_idx, column=21).value)  # U
        filter_electric_heat  = to_str(ws.cell(row=row_idx, column=22).value)  # V

        # Document filenames (basename without extension)
        data_sheet_name = to_str(ws.cell(row=row_idx, column=24).value)  # X
        iom_name        = to_str(ws.cell(row=row_idx, column=25).value)  # Y

        nom_tons_str = format_tons(nom_tons_raw)
        eh_num       = format_heat_int(electric_heat)
        eh_str       = str(eh_num)

        sys_id = f"marvair_{model}_{eh_str}kw"

        systems.append({
            "id": sys_id,
            "schedule": {
                "manufacturer":      "Marvair",
                "model":             model,
                "nomTons":           nom_tons_str,
                "cfm":               cfm,
                "totalCapacity":     total_cap,
                "sensibleCapacity":  sens_cap,
                "outsideAir":        outside_air,
                "enteringAir":       entering_air,
                "electricHeat":      eh_num,
                "voltage":           voltage,
                "mca":               mca,
                "mocp":              mocp,
                "refrigerant":       refrigerant,
                "eer":               eer,
                "ieer":              ieer,
                "configuration":     configuration,
                "ventilation":       ventilation,
            },
            "filters": {
                "size":         filter_tons,
                "electrical":   filter_electrical,
                "electricHeat": filter_electric_heat,
            },
            "docs": {
                "dataSheet": (ASSET_BASE + SUBMITTAL_DIR + data_sheet_name + ".pdf") if data_sheet_name else "",
                "iom":       (ASSET_BASE + IOM_DIR       + iom_name        + ".pdf") if iom_name        else "",
            },
        })

        size_set.add(filter_tons)
        electrical_set.add(filter_electrical)
        electric_heat_set.add(filter_electric_heat)

    # -----------------------------------------------------------------------
    # Part 2 — Parse schedule notes (rows 44-75)
    #
    # Layout is irregular:
    #   Row 44:     "COMMON TO ALL UNITS:"          (section header)
    #   Rows 45-51: Items 1-7 in cols B/C (number/text) and 8-14 in H/I
    #   Row 52:     "OPTIONS FOR OUTSIDE AIR VENTILATION:"  (section header)
    #   Rows 53-56: Items 1-4 in cols B/C
    #   Rows 54-55: "CONFIGURATION:" + items 1-2 in cols K/L (sub-section)
    #   Row 57:     "OPTIONAL ACCESSORIES:"         (section header)
    #   Rows 58-75: Numbered items (B/C) with sub-bullets in D under parent
    #
    # Notes are stored in READING ORDER. Embedded "N-" prefixes are kept
    # intact so preserveNoteNumbering can render them verbatim.
    # -----------------------------------------------------------------------
    notes = []

    # COMMON TO ALL UNITS (14 items split 7+7 across two columns)
    notes.append("COMMON TO ALL UNITS:")
    left_items, right_items = [], []
    for r in range(45, 52):
        num_l  = to_str(ws.cell(row=r, column=2).value)   # B
        text_l = to_str(ws.cell(row=r, column=3).value)   # C
        num_r  = to_str(ws.cell(row=r, column=8).value)   # H
        text_r = to_str(ws.cell(row=r, column=9).value)   # I
        if num_l and text_l:
            left_items.append(f"{num_l} {text_l}")
        if num_r and text_r:
            right_items.append(f"{num_r} {text_r}")
    notes.extend(left_items)
    notes.extend(right_items)

    # OPTIONS FOR OUTSIDE AIR VENTILATION (4 items)
    notes.append("OPTIONS FOR OUTSIDE AIR VENTILATION:")
    for r in range(53, 57):
        num  = to_str(ws.cell(row=r, column=2).value)
        text = to_str(ws.cell(row=r, column=3).value)
        if num and text:
            notes.append(f"{num} {text}")

    # CONFIGURATION sub-section (sits in cols K/L on rows 53-55)
    notes.append("CONFIGURATION:")
    for r in range(54, 56):
        num  = to_str(ws.cell(row=r, column=11).value)   # K
        text = to_str(ws.cell(row=r, column=12).value)   # L
        if num and text:
            notes.append(f"{num} {text}")

    # OPTIONAL ACCESSORIES with sub-bullets
    notes.append("OPTIONAL ACCESSORIES:")
    for r in range(58, 76):
        num      = to_str(ws.cell(row=r, column=2).value)  # B
        text     = to_str(ws.cell(row=r, column=3).value)  # C
        sub_text = to_str(ws.cell(row=r, column=4).value)  # D

        if num and text:
            notes.append(f"{num} {text}")
        elif text == "-" and sub_text:
            # Sub-bullet under the most recent parent item
            notes.append(f"    \u2022 {sub_text}")   # • bullet

    # -----------------------------------------------------------------------
    # Part 3 — Assemble the JSON structure & write
    # -----------------------------------------------------------------------
    data = {
        "productType":                  "Vertical Wall Mount Air Conditioner",
        "manufacturer":                 "Marvair",
        "productKey":                   "marvair-vertical",
        "preserveNoteNumbering":        True,
        "scheduleNotesDefaultChecked":  True,
        "systems":                      systems,
        "filterOptions": {
            "size":         sorted(size_set, key=numkey),
            "electrical":   sorted(electrical_set),
            "electricHeat": sorted(electric_heat_set, key=numkey),
        },
        "scheduleNotes":                notes,
    }

    # Ensure output directory exists
    out_dir = os.path.dirname(OUTPUT_JSON)
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir)

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    # -----------------------------------------------------------------------
    # Summary report
    # -----------------------------------------------------------------------
    print(f"\nWrote {OUTPUT_JSON}")
    print(f"  Systems:           {len(systems)}")
    print(f"  Filter — sizes:    {data['filterOptions']['size']}")
    print(f"  Filter — electr.:  {data['filterOptions']['electrical']}")
    print(f"  Filter — heat kW:  {data['filterOptions']['electricHeat']}")
    print(f"  Schedule notes:    {len(notes)}")


if __name__ == "__main__":
    main()